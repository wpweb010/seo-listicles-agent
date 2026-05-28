#!/usr/bin/env python3
"""
SEO Listicles Agent
===================
Given one keyword + region, fetches organic SERP results (no ads, no sponsored),
filters for pages that LIST companies, validates each URL is live and genuinely
a list, calculates keyword density, optionally finds contact emails, and exports
to Excel.

CLI usage
---------
    python seo_listicles_agent.py "best wordpress development companies"
    python seo_listicles_agent.py --pages 2 --region uk "hire wordpress developers"
    python seo_listicles_agent.py --mode semrush --pages 2 --contacts "top wordpress agencies"

API usage (imported by api.py)
-------------------------------
    from seo_listicles_agent import run, write_excel_history, write_excel_company

Modes
-----
    free     Google Custom Search (100 results max / free tier) + Open PageRank authority
    semrush  Semrush phrase_organic (up to 200 organic results) + Semrush Authority Score

Credentials — set in .env or as environment variables:
    GOOGLE_CSE_KEY, GOOGLE_CSE_CX, OPENPAGERANK_KEY  (free mode)
    SEMRUSH_API_KEY                                   (semrush mode)
"""

import argparse
import os
import re
import sys
import time
import urllib.parse
from datetime import datetime

import requests

# Force UTF-8 output on Windows so log symbols don't crash charmap
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass


# ── Constants ─────────────────────────────────────────────────────────────────

REQUEST_DELAY = 0.4
FETCH_TIMEOUT = 12
CONTACT_PAGES = 1   # extra pages to check when email not found on target

# Region → Google CSE gl code + Semrush database code
REGION_MAP = {
    "us": {"label": "United States",    "cse_gl": "us", "semrush_db": "us"},
    "uk": {"label": "United Kingdom",   "cse_gl": "gb", "semrush_db": "uk"},
    "ca": {"label": "Canada",           "cse_gl": "ca", "semrush_db": "ca"},
    "au": {"label": "Australia",        "cse_gl": "au", "semrush_db": "au"},
    "in": {"label": "India",            "cse_gl": "in", "semrush_db": "in"},
    "de": {"label": "Germany",          "cse_gl": "de", "semrush_db": "de"},
    "fr": {"label": "France",           "cse_gl": "fr", "semrush_db": "fr"},
    "sg": {"label": "Singapore",        "cse_gl": "sg", "semrush_db": "sg"},
    "ae": {"label": "UAE",              "cse_gl": "ae", "semrush_db": "ae"},
}

LISTICLE_URL_SIGNALS = [
    "top", "best", "leading", "list", "companies", "agencies", "firms",
    "vendors", "providers", "ranked", "rated", "hire", "developers",
]
BLOG_PATH_SIGNALS = [
    "/blog", "/blogs", "/articles", "/resources", "/insights",
    "/news", "/guides", "/post", "/posts",
]
LISTICLE_TITLE_SIGNALS = [
    "top ", "best ", "leading ", "list of", "companies", "agencies",
    "firms", "vendors", "providers", "ranked", "rated", "hire",
]
DIRECTORY_DOMAINS = [
    # Major B2B directories
    "clutch.co", "goodfirms.co", "g2.com", "capterra.com", "trustradius.com",
    "sortlist.com", "topdevelopers.co", "themanifest.com", "expertise.com",
    "designrush.com", "appfutura.com", "itfirms.co", "selectedfirms.co",
    "upcity.com", "gartner.com", "forrester.com",
    # Review & rating platforms
    "trustpilot.com", "yelp.com", "yellowpages.com", "bbb.org",
    "glassdoor.com", "ambitionbox.com",
    # Industry-specific directories
    "techreviewer.co", "crowdreviews.com", "businessofapps.com",
    "softwareworld.co", "mobileappdaily.com", "10seos.com",
    # Additional aggregators
    "indiamart.com", "alibaba.com", "upwork.com", "fiverr.com",
    "freelancer.com", "guru.com", "toptal.com", "gun.io",
    "stackoverflow.com", "github.com", "producthunt.com",
]
TARGET_TYPES = (
    "Listicle (blog post)",
    "Listicle (Top/Best)",
    "Blog / Article",
    "Possible listicle",
)
AGGREGATOR_TYPE   = "Aggregator"
SERVICE_PAGE_TYPE = "Service Page"
VIDEO_TYPE        = "Video"
VIDEO_DOMAINS     = ["youtube.com", "youtu.be", "vimeo.com", "tiktok.com",
                     "dailymotion.com", "twitch.tv", "rumble.com"]

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate",  # NO brotli — requests doesn't decode it by default
    "Cache-Control": "no-cache",
    "Pragma": "no-cache",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Connection": "keep-alive",
}
EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
EMAIL_BLOCKLIST = (
    "example.com", "sentry", "wixpress", "@2x", ".png", ".jpg",
    "domain.com", "email.com", "yourdomain",
)
SUBMIT_SIGNALS  = ["write for us", "write-for-us", "contribute", "guest post",
                   "guest-post", "submit", "suggest a", "add your", "get listed"]
CONTACT_SIGNALS = ["contact", "/about", "about-us", "reach-us"]


# ── Helpers ───────────────────────────────────────────────────────────────────

def domain_of(url: str) -> str:
    try:
        netloc = urllib.parse.urlparse(url).netloc.lower()
        return netloc[4:] if netloc.startswith("www.") else netloc
    except Exception:
        return ""


def to_int(v) -> int:
    try:
        return int(float(str(v).replace(",", "")))
    except (ValueError, TypeError):
        return 0


# ── Credentials ───────────────────────────────────────────────────────────────

def _db_key(provider: str) -> dict:
    """Get an API key from DB (returns None if missing). Falls back to .env."""
    try:
        from database import get_api_key
        return get_api_key(provider)
    except Exception:
        return None


def _resolve_key(provider: str, env_name: str) -> str:
    """Try DB first, then .env. Returns the api_key string or empty string."""
    row = _db_key(provider)
    if row and row.get("api_key"):
        return row["api_key"]
    return os.environ.get(env_name, "")


def _resolve_extra(provider: str, field: str, env_name: str = None) -> str:
    """Get an extra field (e.g. CSE cx) from DB extra_data or .env fallback."""
    row = _db_key(provider)
    if row:
        extra = row.get("extra_data") or {}
        if extra.get(field):
            return extra[field]
    if env_name:
        return os.environ.get(env_name, "")
    return ""


def get_creds(mode: str) -> dict:
    if mode == "semrush":
        key = _resolve_key("semrush", "SEMRUSH_API_KEY")
        if not key:
            raise ValueError("Semrush API key not configured (Settings → API Keys, or SEMRUSH_API_KEY in .env).")
        return {"mode": "semrush", "semrush": key}

    # Free mode: prefer Serper.dev, fall back to Google CSE
    serper_key = _resolve_key("serper", "SERPER_API_KEY")
    if serper_key:
        return {
            "mode":     "free",
            "provider": "serper",
            "serper":   serper_key,
            "opr_key":  _resolve_key("openpagerank", "OPENPAGERANK_KEY"),
        }
    cse_key = _resolve_key("google_cse", "GOOGLE_CSE_KEY")
    cse_cx  = _resolve_extra("google_cse", "cx", "GOOGLE_CSE_CX")
    if not cse_key or not cse_cx:
        raise ValueError(
            "No SERP credentials configured. Open Settings → API Keys and add a "
            "Serper.dev key (recommended) OR Google CSE key + Search Engine ID."
        )
    return {
        "mode":     "free",
        "provider": "cse",
        "cse_key":  cse_key,
        "cse_cx":   cse_cx,
        "opr_key":  _resolve_key("openpagerank", "OPENPAGERANK_KEY"),
    }


def _log_api(provider: str, success: bool = True, credits: int = None, meta: dict = None):
    """Log an API call (best-effort)."""
    try:
        from database import log_api_call
        log_api_call(provider, success=success, credits_remaining=credits, metadata=meta)
    except Exception:
        pass


# ── SERP fetching (organic only — no ads, no sponsored) ───────────────────────

def fetch_serp_page(keyword: str, creds: dict, page_num: int, region: str) -> list:
    """
    Fetch exactly one page (10 results) for the given page number.
    Dispatches to the right provider based on creds.
    Returns [] when no more results are available.
    """
    provider = creds.get("provider", "cse")
    mode     = creds.get("mode", "free")
    gl       = REGION_MAP.get(region, REGION_MAP["us"])["cse_gl"]

    # ── Serper.dev ────────────────────────────────────────────────────────────
    if provider == "serper":
        try:
            r = requests.post(
                "https://google.serper.dev/search",
                headers={"X-API-KEY": creds["serper"], "Content-Type": "application/json"},
                json={"q": keyword, "gl": gl, "hl": "en", "num": 10, "page": page_num},
                timeout=20,
            )
            data = r.json()
        except requests.RequestException as e:
            _log_api("serper", success=False, meta={"error": str(e)})
            raise RuntimeError(f"Serper.dev network error: {e}")
        if "error" in data:
            _log_api("serper", success=False, meta={"error": data.get("error")})
            raise RuntimeError(f"Serper.dev error: {data.get('error', data)}")
        # Capture credits if Serper returns them in the response
        credits = data.get("credits") or data.get("credits_remaining")
        _log_api("serper", success=True, credits=credits, meta={"page": page_num})
        organic = data.get("organic", [])
        results = []
        for i, item in enumerate(organic):
            url = item.get("link", "")
            results.append({
                "position": (page_num - 1) * 10 + i + 1,
                "domain":   domain_of(url),
                "url":      url,
                "title":    item.get("title", ""),
                "snippet":  item.get("snippet", ""),
            })
        return results

    # ── Google CSE ────────────────────────────────────────────────────────────
    if provider == "cse":
        start = (page_num - 1) * 10 + 1
        if start > 100:
            return []   # CSE free tier hard cap
        try:
            r = requests.get(
                "https://www.googleapis.com/customsearch/v1",
                params={"key": creds["cse_key"], "cx": creds["cse_cx"],
                        "q": keyword, "num": 10, "start": start,
                        "gl": gl, "hl": "en"},
                timeout=20,
            )
            data = r.json()
        except requests.RequestException as e:
            _log_api("google_cse", success=False, meta={"error": str(e)})
            raise RuntimeError(f"Google CSE network error: {e}")
        if "error" in data:
            msg = data["error"].get("message", "")
            _log_api("google_cse", success=False, meta={"error": msg})
            if "blocked" in msg.lower() or "access" in msg.lower():
                raise RuntimeError(
                    "Custom Search API blocked. Enable billing at console.cloud.google.com "
                    "or switch to Serper.dev (Settings → API Keys)."
                )
            raise RuntimeError(f"Google CSE error: {msg[:200]}")
        _log_api("google_cse", success=True, meta={"page": page_num})
        items = data.get("items", [])
        results = []
        for i, it in enumerate(items):
            url = it.get("link", "")
            results.append({
                "position": start + i,
                "domain":   domain_of(url),
                "url":      url,
                "title":    it.get("title", ""),
                "snippet":  it.get("snippet", ""),
            })
        return results

    # ── Semrush ───────────────────────────────────────────────────────────────
    db     = REGION_MAP.get(region, REGION_MAP["us"])["semrush_db"]
    offset = (page_num - 1) * 10
    params = {
        "type": "phrase_organic", "key": creds["semrush"],
        "phrase": keyword, "database": db,
        "display_offset": offset, "display_limit": 10,
        "export_columns": "Dn,Ur",
    }
    try:
        r    = requests.get("https://api.semrush.com/?" + urllib.parse.urlencode(params), timeout=30)
        text = r.text.strip()
    except requests.RequestException as e:
        raise RuntimeError(f"Semrush network error: {e}")
    if not text or text.startswith("ERROR") or ";" not in text.split("\n")[0]:
        return []
    lines   = [ln for ln in text.split("\n") if ln.strip()]
    headers = [h.strip() for h in lines[0].split(";")]
    results = []
    for i, ln in enumerate(lines[1:]):
        vals = ln.split(";")
        row  = {headers[j]: vals[j].strip() if j < len(vals) else "" for j in range(len(headers))}
        url  = row.get("Url", "").strip()
        results.append({
            "position": offset + i + 1,
            "domain":   row.get("Domain", domain_of(url)).strip(),
            "url":      url, "title": "", "snippet": "",
        })
    return results


# ── Authority ─────────────────────────────────────────────────────────────────

def fetch_openpagerank(domains: list, opr_key: str) -> dict:
    """Open PageRank bulk lookup — free authority proxy, 0-100 scaled."""
    result = {}
    if not opr_key:
        return result
    domains = list(dict.fromkeys(domains))
    for i in range(0, len(domains), 100):
        batch = domains[i:i + 100]
        try:
            r = requests.get(
                "https://openpagerank.com/api/v1.0/getPageRank",
                params=[("domains[]", d) for d in batch],
                headers={"API-OPR": opr_key},
                timeout=20,
            )
            data = r.json()
            for row in data.get("response", []):
                d = row.get("domain", "")
                rank = row.get("page_rank_decimal", 0) or 0
                try:
                    result[d] = round(float(rank) * 10, 1)
                except (ValueError, TypeError):
                    result[d] = ""
            _log_api("openpagerank", success=True, meta={"batch": len(batch)})
        except requests.RequestException as e:
            _log_api("openpagerank", success=False, meta={"error": str(e)})
        time.sleep(REQUEST_DELAY)
    return result


def fetch_semrush_authority(domain: str, key: str, cache: dict) -> str:
    """Semrush domain_ranks — Authority Score (0-100) for one domain."""
    if domain in cache:
        return cache[domain]
    params = {"type": "domain_ranks", "key": key, "domain": domain,
              "database": "us", "export_columns": "Dn,Sh"}
    try:
        r    = requests.get("https://api.semrush.com/?" + urllib.parse.urlencode(params), timeout=30)
        text = r.text.strip()
    except requests.RequestException:
        cache[domain] = ""; return ""
    if not text or text.startswith("ERROR") or ";" not in text.split("\n")[0]:
        cache[domain] = ""; return ""
    lines = [ln for ln in text.split("\n") if ln.strip()]
    if len(lines) < 2:
        cache[domain] = ""; return ""
    headers = [h.strip() for h in lines[0].split(";")]
    vals    = lines[1].split(";")
    row     = {headers[i]: vals[i].strip() if i < len(vals) else "" for i in range(len(headers))}
    score   = row.get("Authority Score", row.get("Sh", ""))
    cache[domain] = score
    time.sleep(REQUEST_DELAY)
    return score


# ── Classification ─────────────────────────────────────────────────────────────

def _get_aggregator_list():
    """Lazy load aggregator list from DB; fall back to hardcoded default."""
    try:
        from database import get_aggregator_domain_list
        domains = get_aggregator_domain_list()
        if domains:
            return domains
    except Exception:
        pass
    return DIRECTORY_DOMAINS


def _get_video_list():
    """Lazy load video domains from DB; fall back to hardcoded default."""
    try:
        from database import get_video_domain_list
        domains = get_video_domain_list()
        if domains:
            return domains
    except Exception:
        pass
    return VIDEO_DOMAINS


def classify(url: str, domain: str, title: str = "") -> str:
    u = (url   or "").lower()
    d = (domain or "").lower()
    t = (title  or "").lower()

    aggregators = _get_aggregator_list()
    videos      = _get_video_list()

    # 1. Aggregator check (highest priority — directory domains)
    if any(dd in d for dd in aggregators):
        return "Aggregator"

    # 1b. Video platforms — YouTube, Vimeo, etc.
    if any(vd in d for vd in videos):
        return VIDEO_TYPE

    # 2. Service Page check — these sell the OWN company's services
    # Strong URL signals that override listicle title hype like "Top WordPress Agency"
    service_page_paths = [
        "/services/", "/service/", "/solutions/",
        "/wordpress-development", "/php-development", "/laravel-development",
        "/web-development", "/mobile-app-development", "/app-development",
        "-development-company", "-development-services", "-development-agency",
        "/hire-", "/hire/",
    ]
    if any(p in u for p in service_page_paths):
        # But override if URL also has strong listicle markers (year, top-N, list-of)
        strong_listicle_url = bool(re.search(r"(top|best|leading)[-_]?\d+", u)) or \
                              bool(re.search(r"\b(202[3-9]|2030)\b", u)) or \
                              "/blog/" in u or "/blogs/" in u or "/articles/" in u
        if not strong_listicle_url:
            return "Service Page"

    # 3. Listicle signals
    url_hits   = sum(1 for s in LISTICLE_URL_SIGNALS   if s in u)
    title_hits = sum(1 for s in LISTICLE_TITLE_SIGNALS if s in t)
    on_blog    = any(seg in u for seg in BLOG_PATH_SIGNALS)

    if url_hits >= 2:
        return "Listicle (blog post)" if on_blog else "Listicle (Top/Best)"
    if url_hits == 1 and on_blog:
        return "Listicle (blog post)"
    if on_blog and title_hits >= 1:
        return "Listicle (blog post)"
    if on_blog:
        return "Blog / Article"
    if url_hits == 1 or title_hits >= 1:
        return "Possible listicle"

    # 4. Homepage / Other
    path = u.split(d, 1)[-1] if d in u else u
    path = path.split("?")[0].rstrip("/")
    if path in ("", "/home", "/index.html"):
        return "Competitor (homepage)"

    return "Competitor / Other"


def priority_score(authority, page_type: str, position: int) -> float:
    a = to_int(authority)
    type_pts = {"Listicle (blog post)": 30, "Listicle (Top/Best)": 25,
                "Blog / Article": 15, "Possible listicle": 8}.get(page_type, 0)
    pos_pts = max(0, 20 - position // 5)
    return round(min(100, a * 0.5 + type_pts + pos_pts), 1)


# ── URL validation + keyword density ─────────────────────────────────────────

def fetch_html(url: str) -> tuple:
    """Returns (status_code, html_text). status=0 on network error.

    Special handling:
    - Reddit URLs: use .json API to get post content (Reddit blocks scrapers).
    """
    # Reddit special case — fetch via JSON API with proper Reddit-API-compliant User-Agent
    if "reddit.com/r/" in url and "/comments/" in url:
        try:
            json_url = url.rstrip("/") + ".json"
            reddit_headers = {
                "User-Agent": "seo-listicles-agent:v1.0 (research tool)",
                "Accept": "application/json",
                "Accept-Language": "en-US,en;q=0.9",
            }
            r = requests.get(json_url, headers=reddit_headers, timeout=FETCH_TIMEOUT,
                             allow_redirects=True)

            # If we got valid JSON, parse it
            if r.status_code == 200 and r.text.strip().startswith(("[", "{")):
                import json as _json
                try:
                    data = _json.loads(r.text)
                    texts = []

                    def walk(obj):
                        if isinstance(obj, dict):
                            for k, v in obj.items():
                                if k in ("selftext", "selftext_html", "title", "body", "body_html"):
                                    if isinstance(v, str) and v:
                                        texts.append(v)
                                else:
                                    walk(v)
                        elif isinstance(obj, list):
                            for item in obj:
                                walk(item)

                    walk(data)
                    combined = "\n".join(texts)
                    import html as _html_mod
                    combined = _html_mod.unescape(combined)
                    if combined.strip():
                        return r.status_code, f"<html><body><div class='reddit-content'>{combined}</div></body></html>"
                except (ValueError, KeyError):
                    pass

            # Fallback: try old.reddit.com which is more scraper-friendly
            try:
                old_url = url.replace("www.reddit.com", "old.reddit.com").replace("reddit.com", "old.reddit.com")
                # Avoid double-replacement
                if "old.reddit.com" not in old_url:
                    old_url = "https://old.reddit.com" + url.split("reddit.com")[-1]
                r2 = requests.get(old_url, headers=reddit_headers, timeout=FETCH_TIMEOUT)
                if r2.status_code == 200 and len(r2.text) > 10000:
                    return r2.status_code, r2.text
            except requests.RequestException:
                pass
        except requests.RequestException:
            pass

    try:
        r = requests.get(url, headers=HEADERS, timeout=FETCH_TIMEOUT, allow_redirects=True)
        html = r.text if "text/html" in r.headers.get("content-type", "") else ""
        return r.status_code, html
    except requests.RequestException:
        return 0, ""


def validate_listicle(url: str) -> tuple:
    """
    Returns (status_code, live, has_list_content, html).

    has_list_content = True when the page genuinely lists multiple companies:
      - has an ordered list OR 5+ list items
      - has numbered headings (e.g. "1. CompanyName")
      - mentions company-type words 3+ times
      - OR has 5+ distinct external domains in <a href> links (suggests company list)
    """
    status, html = fetch_html(url)
    live         = status == 200
    has_list     = False
    if live and html:
        lower = html.lower()
        # Count distinct external domains in links (helps detect service pages listing competitors)
        external_links = re.findall(r'<a[^>]+href=["\']https?://([^/"\']+)', html, re.I)
        external_domains = set()
        page_domain = domain_of(url)
        for link_domain in external_links:
            d = link_domain.lower().replace("www.", "")
            if d != page_domain and not any(skip in d for skip in
                ("facebook.com", "twitter.com", "linkedin.com", "instagram.com",
                 "youtube.com", "google.com", "fonts.googleapis.com", "ajax.googleapis.com",
                 "gravatar.com", "wp.com", "wordpress.org", "cdnjs.com", "jsdelivr.net")):
                external_domains.add(d)

        signals = [
            bool(re.search(r"<ol[^>]*>.*?</ol>", html, re.I | re.S)),
            lower.count("</li>") >= 5,
            bool(re.search(r"<h[1-4][^>]*>\s*\d+[\.\)]", html, re.I)),  # numbered headings
            bool(re.search(r"\b\d+\.\s+[A-Z][a-zA-Z]", html)),  # numbered text patterns
            sum(lower.count(w) for w in
                ("company", "agency", "agencies", "firm", "provider", "vendor")) >= 5,
            len(external_domains) >= 5,  # service pages listing 5+ competitors
        ]
        has_list = sum(signals) >= 2
    return status, live, has_list, html


def keyword_density(html: str, keyword: str) -> float:
    """
    Percentage of words on the page matching the keyword phrase.
    Strips HTML tags first; counts exact phrase occurrences.
    Returns 0.0 if HTML is empty or keyword not found.
    """
    text = re.sub(r"<[^>]+>", " ", html or "")
    text = re.sub(r"\s+", " ", text).lower().strip()
    if not text:
        return 0.0
    words       = text.split()
    total_words = len(words)
    if total_words == 0:
        return 0.0
    kw_clean = keyword.lower().strip()
    count    = len(re.findall(re.escape(kw_clean), text))
    return round((count / total_words) * 100, 2)


# ── Domain detection on listicle pages ────────────────────────────────────

def verify_url_diagnostic(url: str, domain: str) -> dict:
    """
    Diagnostic verification: fetch URL, search for domain, return detailed report.
    Used by /api/verify-url endpoint for user-facing audit trail.
    """
    status_code, html = fetch_html(url)
    if status_code == 0:
        return {
            "url": url, "domain": domain,
            "status": "fetch_error", "http_code": 0,
            "html_size": 0, "variants_found": [],
            "headings": [], "detection": None,
            "summary": "Failed to fetch — network error or timeout",
        }

    # Normalize domain
    domain_name = domain.lower().replace("http://", "").replace("https://", "").split("/")[0]
    domain_name = domain_name.replace("www.", "")
    domain_short = domain_name.split(".")[0]

    # Search for ALL variants
    variants = [
        domain_name, domain_short,
        domain_short.replace("infotech", " infotech") if "infotech" in domain_short else None,
        domain_short.replace("-", " "),
        domain_short.replace("_", " "),
    ]
    variants = [v for v in variants if v]  # Filter None

    variants_found = []
    for v in set(variants):
        count = len(re.findall(re.escape(v), html, re.I))
        if count > 0:
            variants_found.append({"variant": v, "count": count})

    # Compact-text match (no spaces/punct)
    compact = re.sub(r'[\s\-_\.]+', '', html.lower())
    compact_count = compact.count(domain_short)
    if compact_count > 0:
        variants_found.append({"variant": f"(compact: {domain_short})", "count": compact_count})

    # Extract headings (first 20)
    h_pattern = re.compile(r'<(h[1-6])[^>]*>(.*?)</\1>', re.I | re.S)
    headings = []
    for tag, content in h_pattern.findall(html)[:20]:
        clean = re.sub(r'<[^>]+>', '', content).strip()[:100]
        contains_domain = domain_short.lower() in clean.lower().replace(" ", "").replace("-", "")
        headings.append({
            "tag": tag.upper(), "text": clean,
            "contains_domain": contains_domain,
        })

    # Detection result
    detection = find_domain_on_page(domain, html, url)

    # Build human-readable summary
    if detection.get("found"):
        summary = f"LISTED — Found at position {detection['position']} ({detection['link_type']})"
    elif detection.get("link_type") == "bot_protected":
        summary = "BOT-PROTECTED — Page blocked by anti-scraper. Verify manually."
    elif variants_found:
        summary = f"FOUND IN HTML but position unclear (found {sum(v['count'] for v in variants_found)} mentions)"
    else:
        summary = f"NOT LISTED — Domain genuinely absent from HTML (verified {len(html):,} bytes)"

    return {
        "url": url, "domain": domain,
        "status": "success", "http_code": status_code,
        "html_size": len(html),
        "variants_searched": list(set(variants)) + [f"(compact: {domain_short})"],
        "variants_found": variants_found,
        "headings": headings,
        "detection": detection,
        "summary": summary,
    }


def find_domain_on_page(domain: str, html: str, url: str, name_variants: list = None) -> dict:
    """
    Search for a domain/company name on a listicle page.
    Matches BOTH with and without spaces (so "wpwebinfotech" finds "WPWEB Infotech").
    Uses word boundary on domain matching to avoid false positives.

    Optional `name_variants` list: additional company name variants to search for
    (e.g. ["WPWeb Infotech", "WPWeb"]). Improves detection of non-backlinked mentions.

    Returns: {
        "found": bool,
        "position": int or None or "?",
        "link_type": str,    # "Backlink", "Mention", or "—"
    }
    Special: If page is bot-protected (Reddit, Cloudflare wall, etc.), returns
    found=False with link_type="bot_protected" so caller can mark "Check manually".
    """
    if not domain or not html:
        return {"found": False, "position": None, "link_type": "—"}

    # Detect bot-protection / captcha walls
    bot_signals = [
        "please wait for verification",
        "cf-please-wait",
        "checking your browser",
        "please enable javascript and cookies",
        "captcha",
        "rate limit exceeded",
        "access denied",
    ]
    html_lower_head = html[:5000].lower()
    if len(html) < 15000 and any(s in html_lower_head for s in bot_signals):
        return {"found": False, "position": None, "link_type": "bot_protected"}

    # Clean: "wpwebinfotech.com" → domain_name="wpwebinfotech.com", domain_short="wpwebinfotech"
    domain_name = domain.lower().replace("http://", "").replace("https://", "").split("/")[0]
    domain_name = domain_name.replace("www.", "")
    domain_short = domain_name.split(".")[0]

    # Strip HTML & normalize: lowercase, no spaces, no punctuation
    # This lets "wpwebinfotech" match "WPWEB Infotech", "WP Web Infotech", "WPWEB-Infotech"
    plain_text = re.sub(r'<[^>]+>', ' ', html).lower()
    plain_text = re.sub(r'\s+', ' ', plain_text)
    compact_text = re.sub(r'[\s\-_\.\,\:\;]+', '', plain_text)  # No spaces/punct

    # Build search variants — domain + short + any user-provided custom variants
    all_variants = {domain_name, domain_short}
    if name_variants:
        for v in name_variants:
            if v and v.strip():
                all_variants.add(v.strip().lower())
    all_variants = [v for v in all_variants if v]

    # Compile regex patterns for each variant
    variant_patterns = [re.compile(r'\b' + re.escape(v) + r'\b', re.I) for v in all_variants]

    def text_contains(text: str) -> bool:
        """Check using regex word boundaries AND compact (no-spaces) match for each variant."""
        if not text:
            return False
        text_lower = text.lower()
        # Try regex match for each variant
        for pat in variant_patterns:
            if pat.search(text):
                return True
        # Compact match: strip spaces and punctuation
        compact = re.sub(r'[\s\-_\.\,\:\;]+', '', text_lower)
        for v in all_variants:
            v_compact = re.sub(r'[\s\-_\.]+', '', v.lower())
            if v_compact and v_compact in compact:
                return True
        return False

    html_lower = html.lower()

    # Quick existence check
    has_match = False
    for pat in variant_patterns:
        if pat.search(html_lower):
            has_match = True
            break
    if not has_match:
        # Try compact match
        for v in all_variants:
            v_compact = re.sub(r'[\s\-_\.]+', '', v.lower())
            if v_compact and v_compact in compact_text:
                has_match = True
                break
    if not has_match:
        return {"found": False, "position": None, "link_type": "—"}

    # Determine link type — is the FULL domain wrapped in an <a href> tag?
    has_backlink = bool(re.search(
        r'<a[^>]+href=["\'][^"\']*\b' + re.escape(domain_name) + r'\b[^"\']*["\']',
        html, re.I
    ))
    link_type = "Backlink" if has_backlink else "Mention"

    # Strategy 1: Numbered headings like "1. Company Name" (handles nested tags)
    h_pattern = re.compile(r'<(h[1-6])[^>]*>(.*?)</\1>', re.I | re.S)
    all_h_blocks = h_pattern.findall(html)  # [(tag, content), ...]

    numbered_headings = []  # (number, clean_text)
    headings_by_level = {1: [], 2: [], 3: [], 4: [], 5: [], 6: []}
    all_headings = []  # (level, clean_text) in order

    for tag, content in all_h_blocks:
        clean = re.sub(r'<[^>]+>', '', content).strip()
        level = int(tag[1])
        headings_by_level[level].append(clean)
        all_headings.append((level, clean))
        m = re.match(r'^\s*(\d{1,2})[\.\)\:\s]+(.*)', clean)
        if m:
            num = int(m.group(1))
            rest = m.group(2)
            numbered_headings.append((num, rest, clean))

    # Strategy 1: Numbered headings — most reliable (e.g., LinkedIn's "1. WPWeb Infotech")
    for num, rest, full in numbered_headings:
        if text_contains(rest) or text_contains(full):
            return {"found": True, "position": num, "link_type": link_type}

    # Strategy 2: Find the heading level used for the COMPANY LIST
    # The listicle typically uses H2 or H3 for each company.
    # Find the level with the most headings that are NOT the page title and contain text
    # Skip H1 (almost always page title, not part of the list)
    best_level = None
    best_count = 0
    for level in [2, 3, 4]:
        # Only consider levels with multiple headings (a real list)
        headings = headings_by_level[level]
        if len(headings) >= 3 and len(headings) > best_count:
            best_count = len(headings)
            best_level = level

    if best_level:
        # Count headings AT THAT LEVEL only, in order
        for i, h_text in enumerate(headings_by_level[best_level], 1):
            if text_contains(h_text):
                return {"found": True, "position": i, "link_type": link_type}

    # Strategy 2b: Fall back to counting H2/H3/H4 in order (skip H1 = page title)
    pos = 0
    for level, h_text in all_headings:
        if level == 1:
            continue  # Skip H1 (page title)
        pos += 1
        if text_contains(h_text):
            return {"found": True, "position": pos, "link_type": link_type}

    # Strategy 2c: <p>/<div> with numbered prefix
    p_pattern = re.compile(
        r'<(?:p|div)[^>]*>(.*?)</(?:p|div)>',
        re.I | re.S
    )
    for content in p_pattern.findall(html):
        clean = re.sub(r'<[^>]+>', '', content).strip()
        m = re.match(r'^\s*(\d{1,2})[\.\)\:\s]+(.*)', clean)
        if m:
            num = int(m.group(1))
            rest = m.group(2)[:200]
            if text_contains(rest):
                return {"found": True, "position": num, "link_type": link_type}

    # Strategy 3: Numbered patterns in clean text "1. CompanyName"
    clean_text = re.sub(r'<[^>]+>', ' ', html)
    clean_text = re.sub(r'\s+', ' ', clean_text)
    # Search for "N. ... domain" within 200 chars
    pattern = r'(?:^|\s)(\d{1,3})[\.\)\:][^\d]{0,200}?\b' + re.escape(domain_short) + r'\b'
    for match in re.finditer(pattern, clean_text, re.I):
        try:
            num = int(match.group(1))
            if 1 <= num <= 100:
                return {"found": True, "position": num, "link_type": link_type}
        except (ValueError, IndexError):
            continue

    # Strategy 4: Ordered list <ol> items (not <ul> which often = nav menus)
    ol_match = re.search(r'<ol[^>]*>(.*?)</ol>', html, re.I | re.S)
    if ol_match:
        li_items = re.findall(r'<li[^>]*>(.*?)</li>', ol_match.group(1), re.I | re.S)
        for i, li in enumerate(li_items, 1):
            if text_contains(li):
                return {"found": True, "position": i, "link_type": link_type}

    # Strategy 5: Count external company <a href> links in order they appear
    # Many listicles use "<h2>1. <a href='https://company.com'>Company Name</a></h2>" or similar
    # We extract all external href domains in order and find where our target is
    page_domain_of_listicle = urllib.parse.urlparse(url).netloc.lower().replace("www.", "")
    skip_domains = {
        "facebook.com", "twitter.com", "linkedin.com", "instagram.com", "youtube.com",
        "google.com", "googleapis.com", "gstatic.com", "gravatar.com", "wp.com",
        "wordpress.org", "wordpress.com", "cdnjs.com", "jsdelivr.net", "unpkg.com",
        "cloudflare.com", "doubleclick.net", "googletagmanager.com", "googleadservices.com",
        page_domain_of_listicle,
    }
    seen_companies = []  # Ordered list of unique company domains as they appear
    for m in re.finditer(r'<a[^>]+href=["\']https?://([^/"\']+)', html, re.I):
        link_d = m.group(1).lower().replace("www.", "")
        # Skip social/CDN domains and the listicle's own domain
        if any(skip in link_d for skip in skip_domains):
            continue
        if link_d not in seen_companies:
            seen_companies.append(link_d)

    # Find our target domain in the ordered list
    for i, comp_d in enumerate(seen_companies, 1):
        if domain_name in comp_d or domain_short in comp_d.split(".")[0]:
            return {"found": True, "position": i, "link_type": link_type}

    # Strategy 6: <strong>/<b> tags hold company names — count them in order
    strong_blocks = re.findall(r'<(?:strong|b)[^>]*>(.*?)</(?:strong|b)>', html, re.I | re.S)
    seen_company_names = []
    for s_content in strong_blocks:
        s_text = re.sub(r'<[^>]+>', '', s_content).strip()
        if len(s_text) >= 3 and len(s_text) <= 60 and not re.match(r'^[\d\$\.]+$', s_text):
            if s_text not in seen_company_names:
                seen_company_names.append(s_text)

    for i, name in enumerate(seen_company_names, 1):
        name_compact = re.sub(r'[\s\-_\.]+', '', name.lower())
        if domain_short in name_compact:
            return {"found": True, "position": i, "link_type": link_type}

    # Strategy 7: Find any "\d+\.\s*<company>" pattern where company contains domain (compact match)
    # This handles "1. WPWEB Infotech" in plain text
    # Look for "N. " patterns anywhere followed by text containing the domain (compact match)
    for match in re.finditer(r'(?:^|\s|>)(\d{1,2})[\.\)]\s+([^\.\n<]{3,80})', plain_text, re.M):
        try:
            num = int(match.group(1))
            content = match.group(2)
            content_compact = re.sub(r'[\s\-_\.]+', '', content)
            if 1 <= num <= 50 and domain_short in content_compact:
                return {"found": True, "position": num, "link_type": link_type}
        except (ValueError, IndexError):
            continue

    # Domain found in text but exact position unclear
    return {"found": True, "position": "?", "link_type": link_type}


# ── Contact detection ─────────────────────────────────────────────────────────

def extract_emails(html: str) -> list:
    found = []
    for m in EMAIL_RE.findall(html or ""):
        e = m.lower()
        if not any(b in e for b in EMAIL_BLOCKLIST) and e not in found:
            found.append(e)
    preferred = [e for e in found if any(
        k in e for k in ("editor", "content", "press", "media", "hello", "info", "contact"))]
    return (preferred + [e for e in found if e not in preferred])[:3]


def find_links(html: str, base_url: str, signals: list) -> list:
    hits = []
    for m in re.finditer(r'<a[^>]+href=["\']([^"\']+)["\'][^>]*>(.*?)</a>',
                         html or "", re.I | re.S):
        href = m.group(1)
        text = re.sub(r"<[^>]+>", " ", m.group(2)).lower()
        if any(s in (href + " " + text).lower() for s in signals):
            abs_href = urllib.parse.urljoin(base_url, href)
            if abs_href not in hits:
                hits.append(abs_href)
    return hits


def detect_email(url: str, domain: str, html: str) -> dict:
    """Find an email address on the target page; falls back to /contact once."""
    result = {"email": "", "submit_url": "", "contact_url": ""}
    emails = extract_emails(html)
    if emails:
        result["email"] = emails[0]
    subs = find_links(html, url, SUBMIT_SIGNALS)
    if subs:
        result["submit_url"] = subs[0]
    cons = find_links(html, url, CONTACT_SIGNALS)
    if cons:
        result["contact_url"] = cons[0]
    if not result["email"] and CONTACT_PAGES > 0:
        fallback = result["contact_url"] or f"https://{domain}/contact"
        _, html2 = fetch_html(fallback)
        if html2:
            em2 = extract_emails(html2)
            if em2:
                result["email"] = em2[0]
            if not result["submit_url"]:
                sl = find_links(html2, fallback, SUBMIT_SIGNALS)
                if sl:
                    result["submit_url"] = sl[0]
        time.sleep(REQUEST_DELAY)
    return result


# ── Excel export ──────────────────────────────────────────────────────────────

def write_excel_company(company_name: str, company_domain: str, rows: list, output):
    """
    Write a per-company Excel report.
    Includes a header row showing the company name + domain.
    """
    return write_excel_history(rows, output, header_context={
        "company_name":   company_name,
        "company_domain": company_domain,
    })


def write_excel_history(rows: list, output, header_context: dict = None):
    """
    Write history export with multi-keyword grouped data.
    Format: Keyword & Location | Domain | Pages | Page Type | Status | Position | Link Type | Notes
    Rows are grouped by keyword. Only Page Type and Status cells are colored.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
    HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
    GROUP_FILL = PatternFill("solid", fgColor="2E75B6")
    GROUP_FONT = Font(color="FFFFFF", bold=True, size=11)
    THIN = Side(style="thin", color="D9D9D9")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    OUTREACH_COLS = ["Keyword & Location", "Domain", "Pages", "Page Type", "Status",
                     "Position", "Link Type", "Notes"]

    def get_page_type_color(page_type):
        """Color for Page Type cell only."""
        if page_type == "Aggregator":
            return PatternFill("solid", fgColor="FFE6CC")  # Light Orange
        if page_type == "Service Page":
            return PatternFill("solid", fgColor="F0E6FF")  # Light Purple
        if page_type == "Video":
            return PatternFill("solid", fgColor="FFD6E0")  # Light Pink
        if page_type and page_type.startswith("Listicle"):
            return PatternFill("solid", fgColor="E7F3FF")  # Light Blue
        if page_type == "Blog / Article":
            return PatternFill("solid", fgColor="E7F3FF")  # Light Blue
        if page_type == "Possible listicle":
            return PatternFill("solid", fgColor="FFF7E6")  # Light Yellow
        return None

    def get_status_color(status):
        """Color for Status cell only."""
        if status == "Listed":
            return PatternFill("solid", fgColor="C6EFCE")  # Green
        if status == "Check Manually":
            return PatternFill("solid", fgColor="FFF2CC")  # Light Yellow
        return None  # Blank cell (no fill) for empty/Not Listed

    ws = wb.active
    ws.title = "Search Results"

    col_widths = {"Keyword & Location": 35, "Domain": 25, "Pages": 50,
                  "Page Type": 22, "Status": 14, "Position": 12,
                  "Link Type": 14, "Notes": 30}

    # Optional context header (when exporting a company report)
    header_offset = 0
    if header_context:
        title = f"Outreach Report — {header_context.get('company_name', '')}"
        if header_context.get('company_domain'):
            title += f" ({header_context['company_domain']})"
        ws.cell(1, 1, title)
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=len(OUTREACH_COLS))
        c1 = ws.cell(1, 1)
        c1.font = Font(bold=True, size=14, color="FFFFFF")
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.fill = PatternFill("solid", fgColor="1F4E78")
        ws.row_dimensions[1].height = 28

        ws.cell(2, 1, f"Generated: {datetime.now():%Y-%m-%d %A %H:%M}")
        ws.merge_cells(start_row=2, start_column=1,
                       end_row=2, end_column=len(OUTREACH_COLS))
        c2 = ws.cell(2, 1)
        c2.font = Font(italic=True, color="5A6A7E", size=10)
        c2.alignment = Alignment(horizontal="center")
        header_offset = 2

    # Column headers
    header_row = header_offset + 1
    for c, col in enumerate(OUTREACH_COLS, 1):
        cell = ws.cell(header_row, c, col)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = col_widths.get(col, 16)
    ws.freeze_panes = f"A{header_row + 1}"

    # Group rows by keyword+location
    grouped = {}
    for r in rows:
        key = f"{r.get('Keyword', '')} ({r.get('Location', 'us').upper()})"
        if key not in grouped:
            grouped[key] = []
        grouped[key].append(r)

    # Write grouped rows — keyword appears ONLY in first row of each group,
    # with an EMPTY ROW separator between groups
    current_row = header_row + 1
    group_index = 0
    total_groups = len(grouped)
    for keyword_loc, group_rows in grouped.items():
        group_rows.sort(key=lambda r: r.get("Position", 9999) or 9999)

        first_row_in_group = current_row

        for idx, r in enumerate(group_rows):
            kw_cell_value = keyword_loc if idx == 0 else ""

            # Suppress "Not Listed" text — leave blank for cleaner reading
            raw_status = r.get("Status", "") or ""
            display_status = "" if raw_status == "Not Listed" else raw_status
            raw_pos = r.get("Position on Page", "") or ""
            display_pos = "" if raw_pos == "—" else raw_pos
            raw_link = r.get("Link Type", "") or ""
            display_link = "" if raw_link == "—" else raw_link

            row_data = [
                kw_cell_value,
                r.get("Domain", "") or "",
                r.get("URL", "") or "",
                r.get("Page Type", "") or "",
                display_status,
                display_pos,
                display_link,
                r.get("Notes", "") or "",
            ]
            page_type = r.get("Page Type", "") or ""
            status = r.get("Status", "") or ""

            for c, val in enumerate(row_data, 1):
                cell = ws.cell(current_row, c, val)
                cell.border = BORDER
                cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Make first row of group BOLD on column A
            if idx == 0:
                ws.cell(current_row, 1).font = Font(bold=True, color="1F4E78")

            # Color ONLY Page Type cell
            pt_color = get_page_type_color(page_type)
            if pt_color:
                ws.cell(current_row, 4).fill = pt_color

            # Color ONLY Status cell
            st_color = get_status_color(status)
            if st_color:
                ws.cell(current_row, 5).fill = st_color

            current_row += 1

        group_index += 1
        # Add EMPTY row separator between groups (not after the last group)
        if group_index < total_groups:
            current_row += 1  # leave a blank row

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(OUTREACH_COLS))}{ws.max_row}"

    wb.save(output)


def write_excel(keyword: str, rows: list, output, region: str = "us"):
    """
    Write results to Excel with two tabs: Listicles | Aggregators.
    Format matches user's outreach template.
    `output` can be a file path string or a BytesIO buffer.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
    HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
    PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
    WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
    THIN      = Side(style="thin", color="D9D9D9")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    # Unified sheet: Keyword & Location | Domain | Pages | Page Type | Status | Position | Link Type | Notes
    OUTREACH_COLS = ["Keyword & Location", "Domain", "Pages", "Page Type", "Status",
                     "Position", "Link Type", "Notes"]

    def col_w(name):
        if "Pages" in name or name == "Notes": return 50
        if "Domain" in name: return 28
        if "Page Type" in name or "Status" in name: return 18
        return 16

    def make_header(ws, cols):
        ws.append(cols)
        for c, col in enumerate(cols, 1):
            cell = ws.cell(1, c)
            cell.fill = HEAD_FILL; cell.font = HEAD_FONT
            cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
            cell.border = BORDER
            ws.column_dimensions[get_column_letter(c)].width = col_w(col)
        ws.freeze_panes = "A2"

    def get_page_type_color(page_type):
        if page_type == "Aggregator":
            return PatternFill("solid", fgColor="FFE6CC")  # Orange
        if page_type and page_type.startswith("Listicle"):
            return PatternFill("solid", fgColor="E7F3FF")  # Blue
        if page_type == "Blog / Article":
            return PatternFill("solid", fgColor="F0E6FF")  # Purple
        if page_type == "Possible listicle":
            return PatternFill("solid", fgColor="FFF7E6")  # Yellow
        return None

    def get_status_color(status):
        if status == "Listed":
            return PatternFill("solid", fgColor="C6EFCE")  # Green
        if status == "Check Manually":
            return PatternFill("solid", fgColor="FFF2CC")  # Light Yellow
        # Not Listed = white background (no fill)
        return None

    def apply_rows(ws, cols, data, keyword="", region=""):
        keyword_location = f"{keyword} ({region.upper()})" if keyword else ""
        for idx, r in enumerate(data):
            row_data = [
                keyword_location if idx == 0 else "",  # Only first row shows keyword
                r.get("Domain", ""),
                r.get("URL", ""),
                r.get("Page Type", ""),
                r.get("Status", ""),
                r.get("Position on Page", ""),
                r.get("Link Type", ""),
                r.get("Notes", ""),
            ]
            ws.append(row_data)

        # Apply formatting — color ONLY Page Type and Status cells
        for ridx in range(2, ws.max_row + 1):
            page_type = ws.cell(ridx, 4).value or ""
            status = ws.cell(ridx, 5).value or ""

            for c in range(1, len(cols) + 1):
                ws.cell(ridx, c).border = BORDER
                ws.cell(ridx, c).alignment = Alignment(wrap_text=True, vertical="top")

            # Make first row's keyword cell bold
            if ridx == 2:
                ws.cell(ridx, 1).font = Font(bold=True, color="1F4E78")

            # Color the Page Type cell (col 4)
            pt_color = get_page_type_color(page_type)
            if pt_color:
                ws.cell(ridx, 4).fill = pt_color

            # Color the Status cell (col 5)
            st_color = get_status_color(status)
            if st_color:
                ws.cell(ridx, 5).fill = st_color
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"

    # Single unified sheet: all results (listicles + aggregators)
    ws_data = wb.active
    ws_data.title = "Search Results"
    all_results = sorted(rows, key=lambda r: r.get("Position", 9999))
    make_header(ws_data, OUTREACH_COLS)
    apply_rows(ws_data, OUTREACH_COLS, all_results, keyword=keyword, region=region)

    wb.save(output)


# ── Main pipeline ─────────────────────────────────────────────────────────────

def run(
    keyword: str,
    domain: str           = None,   # Optional domain to search for on pages
    name_variants: list   = None,   # Optional list of company name variants
    mode: str             = "free",
    target_listicles: int = 10,
    max_pages: int        = 20,
    region: str           = "us",
    find_contacts: bool   = False,
    progress_cb           = None,   # callable({"type": "log"|"done", ...})
    write_file: bool      = True,
    # legacy compat — ignored; use target_listicles instead
    pages: int            = None,
) -> list:
    """
    Run the full pipeline and return a list of result dicts.

    Fetches SERP pages one at a time until `target_listicles` listicle pages
    are found or `max_pages` SERP pages have been searched.
    Directories (Clutch, G2, etc.) are silently dropped — never shown in output.

    If write_file=True, also saves an Excel file.
    progress_cb receives {"type": "log", "msg": "..."} events during the run.
    """
    def log(msg: str):
        print(msg)
        if progress_cb:
            progress_cb({"type": "log", "msg": msg})

    creds  = get_creds(mode)
    region = region if region in REGION_MAP else "us"
    region_label = REGION_MAP[region]["label"]

    log(f"Keyword : {keyword}")
    if domain:
        log(f"Domain  : {domain}")
    log(f"Region  : {region_label}")
    log(f"Target  : {target_listicles} listicle pages")
    log(f"Mode    : {mode}")

    provider      = creds.get("provider", "cse")
    provider_name = ("Serper.dev" if provider == "serper"
                     else "Google CSE" if provider == "cse"
                     else "Semrush")
    log(f"-- Searching via {provider_name} (up to {max_pages} SERP pages)...")

    # -- 1. Smart SERP pagination — track categories: listicles, service pages, aggregators, video
    collected        = []   # listicles only — capped at target_listicles
    aggregators      = []   # aggregators found in the SERP range
    service_pages    = []   # service pages found in the SERP range
    videos           = []   # video platform pages (YouTube etc.)
    listicle_count   = 0
    aggregator_count = 0
    service_count    = 0
    video_count      = 0
    page_num         = 1
    last_serp_pos    = 0

    while listicle_count < target_listicles and page_num <= max_pages:
        try:
            page_results = fetch_serp_page(keyword, creds, page_num, region)
        except RuntimeError as e:
            log(f"  SERP error on page {page_num}: {e}")
            break
        if not page_results:
            log("  No more results from search engine.")
            break

        new_listicles = 0
        for item in page_results:
            if listicle_count >= target_listicles:
                break

            page_type = classify(item["url"], item["domain"], item.get("title", ""))

            if page_type == AGGREGATOR_TYPE:
                aggregators.append((item, page_type))
                aggregator_count += 1
                last_serp_pos = max(last_serp_pos, item["position"])
            elif page_type == SERVICE_PAGE_TYPE:
                service_pages.append((item, page_type))
                service_count += 1
                last_serp_pos = max(last_serp_pos, item["position"])
            elif page_type == VIDEO_TYPE:
                videos.append((item, page_type))
                video_count += 1
                last_serp_pos = max(last_serp_pos, item["position"])
            elif page_type in TARGET_TYPES:
                collected.append((item, page_type))
                listicle_count += 1
                new_listicles  += 1
                last_serp_pos = max(last_serp_pos, item["position"])
            # Skip non-target types (Competitor / Other, Homepage)

        log(f"  Page {page_num}: {len(page_results)} results | "
            f"+{new_listicles} listicles | {listicle_count}/{target_listicles} target listicles found")

        if listicle_count >= target_listicles:
            log(f"  Target reached: {target_listicles} listicles + {service_count} service pages + "
                f"{aggregator_count} aggregators + {video_count} videos")
            break

        page_num += 1
        time.sleep(REQUEST_DELAY)

    total = len(collected) + service_count + aggregator_count + video_count
    log(f"  Output: {len(collected)} listicles + {service_count} service + {aggregator_count} aggr + "
        f"{video_count} video = {total} total rows")

    if not collected and not aggregators:
        raise RuntimeError("No SERP results returned. Check credentials and keyword.")

    # -- 2. Authority scores --------------------------------------------------
    all_items = [(item, pt, "listicle") for item, pt in collected] + \
                [(item, pt, "service_page") for item, pt in service_pages] + \
                [(item, pt, "aggregator") for item, pt in aggregators] + \
                [(item, pt, "video") for item, pt in videos]

    if mode == "free" and creds.get("opr_key"):
        log("-- Fetching Open PageRank authority scores...")
        domains  = list(dict.fromkeys(item["domain"] for item, _, _ in all_items))
        opr_data = fetch_openpagerank(domains, creds["opr_key"])
    else:
        opr_data = {}

    sem_cache = {}

    # -- 3. Build rows --------------------------------------------------------
    rows = []
    for item, page_type, category in all_items:
        if mode == "semrush":
            auth = fetch_semrush_authority(item["domain"], creds["semrush"], sem_cache)
        else:
            auth = opr_data.get(item["domain"], "")
        is_target = page_type in TARGET_TYPES
        rows.append({
            "Position":          item["position"],
            "Domain":            item["domain"],
            "URL":               item["url"],
            "Title":             item.get("title", ""),
            "Page Type":         page_type,
            "Listicle Target?":  "YES" if is_target else "NO",
            "Authority Score":   auth,
            "Priority Score":    priority_score(auth, page_type, item["position"]),
            "Keyword Density %": "",
            "Live?":             "",
            "Has List Content?": "",
            "Validated?":        "",
            "Email":             "",
            "Submit URL":        "",
            "Contact URL":       "",
            "Status":            "",  # Will be populated if domain provided
            "Position on Page":  "",  # Will be populated if domain provided
            "Link Type":         "",  # Will be populated if domain provided
            "Category":          category,  # "listicle" or "aggregator"
        })

    # -- 4. Validate listicles AND service pages (skip aggregators) -----------
    targets = [r for r in rows if r["Category"] in ("listicle", "service_page")]
    log(f"-- Validating {len(targets)} pages (listicles + service pages)...")
    if domain:
        log(f"   (+ searching for {domain}...)")

    html_cache      = {}
    validated_count = 0
    domain_found_count = 0

    for i, r in enumerate(targets, 1):
        log(f"  [{i}/{len(targets)}] {r['Domain']}")
        status, live, has_list, html = validate_listicle(r["URL"])
        r["Live?"]             = live
        r["Has List Content?"] = has_list
        r["Validated?"]        = live and has_list

        # NOTE: Service Pages stay as "Service Page" even if they list competitors.
        # They're a SEPARATE category from listicles per user requirement.

        if live and html:
            r["Keyword Density %"] = keyword_density(html, keyword)
            if find_contacts:
                html_cache[r["URL"]] = html

            # If domain provided, search for it on this page (even if content validation failed)
            if domain:
                detection = find_domain_on_page(domain, html, r["URL"], name_variants=name_variants)
                if detection["found"]:
                    r["Status"] = "Listed"
                    r["Position on Page"] = detection["position"] if detection["position"] is not None else "?"
                    r["Link Type"] = detection["link_type"]
                    domain_found_count += 1
                    log(f"       found: position {detection['position']}, {detection['link_type']}")
                elif detection.get("link_type") == "bot_protected":
                    # Page blocked by anti-bot (Reddit, Cloudflare, etc.) — can't determine
                    r["Status"] = "Check Manually"
                    r["Position on Page"] = "?"
                    r["Link Type"] = "?"
                    r["Notes"] = "Bot-protected page — verify manually"
                    log(f"       bot-protected: check manually")
                else:
                    r["Status"] = "Not Listed"
                    r["Position on Page"] = "—"
                    r["Link Type"] = "—"

        if r["Validated?"]:
            validated_count += 1
            log(f"    [OK] validated")
        elif live:
            log(f"    [~] live, list content not confirmed")
        else:
            code = status if status else "timeout"
            log(f"    [FAIL] dead (HTTP {code})")
        time.sleep(REQUEST_DELAY)

    log(f"  {validated_count}/{len(targets)} fully validated")
    if domain:
        log(f"  {domain_found_count}/{len(targets)} pages list {domain}")

    # -- 5. Find emails -------------------------------------------------------
    if find_contacts:
        live_targets = [r for r in targets if r.get("Live?")]
        log(f"-- Finding emails for {len(live_targets)} live pages...")
        for i, r in enumerate(live_targets, 1):
            html    = html_cache.get(r["URL"], "")
            contact = detect_email(r["URL"], r["Domain"], html)
            r["Email"]       = contact["email"]
            r["Submit URL"]  = contact["submit_url"]
            r["Contact URL"] = contact["contact_url"]
            found = contact["email"] if contact["email"] else "not found"
            log(f"  [{i}/{len(live_targets)}] {r['Domain']}  email: {found}")
    else:
        log("-- Contact detection skipped (enable 'Find email contacts' to turn on)")

    # -- 6. Export ------------------------------------------------------------
    if write_file:
        safe  = re.sub(r"[^a-z0-9]+", "_", keyword.lower()).strip("_")
        stamp = datetime.now().strftime("%Y%m%d_%H%M")
        out   = f"listicles_{safe}_{stamp}.xlsx"
        log(f"-- Saving {out}")
        write_excel(keyword, rows, out, region=region)
        log(f"   Saved: {out}")

    log(f"-- Done: {len(rows)} results | {len(targets)} listicles | {validated_count} validated")
    return rows


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Find Google listicle pages that list companies for a keyword."
    )
    parser.add_argument("keyword", nargs="?",
                        help='Keyword, e.g. "best wordpress development companies"')
    parser.add_argument("--mode", choices=["free", "semrush"], default="free")
    parser.add_argument("--target", type=int, default=10,
                        help="Number of listicle pages to find (default: 10)")
    parser.add_argument("--max-pages", type=int, default=20,
                        help="Max SERP pages to search (default: 20)")
    parser.add_argument("--region", choices=list(REGION_MAP.keys()), default="us",
                        help="Target region (default: us)")
    parser.add_argument("--contacts", action="store_true",
                        help="Find email addresses on listicle pages")
    args    = parser.parse_args()
    keyword = args.keyword or input("Keyword: ").strip()
    if not keyword:
        sys.exit("No keyword provided.")
    try:
        run(keyword, args.mode, args.target, args.max_pages,
            args.region, args.contacts, write_file=True)
    except (ValueError, RuntimeError) as e:
        sys.exit(f"Error: {e}")


if __name__ == "__main__":
    main()
